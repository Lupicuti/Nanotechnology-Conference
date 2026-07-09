import os
import sys
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing via pip...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def create_conference_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # -------------------------------------------------------------
    # 1. CREATE 'Risposte del modulo 1' TAB (Main Data Source)
    # -------------------------------------------------------------
    ws_data = wb.create_sheet(title="Risposte del modulo 1")
    
    headers = [
        "Informazioni cronologiche",
        "Full Name",
        "Email Address",
        "Academic Status / Role",
        "Day 2 Laboratory Workshop",
        "Session Attending"
    ]
    
    # Check if existing CSV exists to import data
    csv_filename = "Sapienza Conference 2026 Registration - Risposte del modulo 1.csv"
    data_rows = []
    if os.path.exists(csv_filename):
        with open(csv_filename, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)
            if len(rows) > 0:
                headers = rows[0]
            if len(rows) > 1:
                data_rows = rows[1:]
    
    ws_data.append(headers)
    for row in data_rows:
        ws_data.append(row)
        
    # Style Data Headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="121A2D", end_color="121A2D", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_data.row_dimensions[1].height = 28
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # -------------------------------------------------------------
    # 2. CREATE 'Dashboard' TAB (Analytics & Counts)
    # -------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Dashboard", index=0) # Put as first tab
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash.merge_cells("A1:E2")
    title_cell = ws_dash["A1"]
    title_cell.value = "SAPIENZA NANO & TECH 2026 — LIVE ANALYTICS DASHBOARD"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Section 1: Session Attendance Breakdown
    ws_dash["A4"] = "PLENARY & WORKSHOP SESSION ATTENDANCE breakdown"
    ws_dash["A4"].font = Font(name="Calibri", size=12, bold=True, color="00F0FF")
    
    ws_dash["A5"] = "Session Name / Time Slot"
    ws_dash["B5"] = "Total Attendees Registered"
    ws_dash["A5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_dash["B5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_dash["A5"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws_dash["B5"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    sessions = [
        ("Session 1: Sept 17 Morning (Plenary Speeches & Optics)", '*Session 1*'),
        ("Session 2: Sept 17 Afternoon (Synthesis & Nanorestoration)", '*Session 2*'),
        ("Session 3: Sept 18 Morning (Reticular Chem & Nanofluidics)", '*Session 3*'),
        ("Session 4: Sept 18 Afternoon (Hands-on Laboratory Workshop)", '*Session 4*')
    ]
    
    for i, (label, pattern) in enumerate(sessions, start=6):
        ws_dash[f"A{i}"] = label
        ws_dash[f"B{i}"] = f'=COUNTIF(\'Risposte del modulo 1\'!F:F, "{pattern}")'
        ws_dash[f"A{i}"].font = Font(name="Calibri", size=11)
        ws_dash[f"B{i}"].font = Font(name="Calibri", size=11, bold=True, color="10B981")
        ws_dash[f"B{i}"].alignment = Alignment(horizontal="center")
        
    # Section 2: Laboratory Workshop Breakdown
    ws_dash["A12"] = "RESERVED LABORATORY WORKSHOP SEATS BREAKDOWN"
    ws_dash["A12"].font = Font(name="Calibri", size=12, bold=True, color="10B981")
    
    ws_dash["A13"] = "Laboratory Workshop Name"
    ws_dash["B13"] = "Seats Reserved"
    ws_dash["A13"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_dash["B13"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws_dash["A13"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws_dash["B13"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    workshops = [
        ("Atomic Force Microscopy (AFM)", '*Atomic Force Microscopy*'),
        ("Nanomedicine Experience", '*Nanomedicine*'),
        ("Absorption in Quantum Well", '*Quantum Well*'),
        ("Semiconductor Nanowires", '*Nanowires*'),
        ("HERITAGE-LAB Nanotech", '*HERITAGE*'),
        ("Light and Photonics", '*Photonics*'),
        ("Computational Plasmonics", '*Computational*'),
        ("None (Plenary Only)", '*Plenary Only*')
    ]
    
    for i, (label, pattern) in enumerate(workshops, start=14):
        ws_dash[f"A{i}"] = label
        ws_dash[f"B{i}"] = f'=COUNTIF(\'Risposte del modulo 1\'!E:E, "{pattern}")'
        ws_dash[f"A{i}"].font = Font(name="Calibri", size=11)
        ws_dash[f"B{i}"].font = Font(name="Calibri", size=11, bold=True, color="00F0FF")
        ws_dash[f"B{i}"].alignment = Alignment(horizontal="center")
        
    ws_dash.column_dimensions["A"].width = 60
    ws_dash.column_dimensions["B"].width = 28
    
    # Thin borders for tables
    thin_border = Border(left=Side(style='thin', color='CBD5E1'),
                         right=Side(style='thin', color='CBD5E1'),
                         top=Side(style='thin', color='CBD5E1'),
                         bottom=Side(style='thin', color='CBD5E1'))
    for r in range(5, 10):
        ws_dash[f"A{r}"].border = thin_border
        ws_dash[f"B{r}"].border = thin_border
    for r in range(13, 22):
        ws_dash[f"A{r}"].border = thin_border
        ws_dash[f"B{r}"].border = thin_border

    # -------------------------------------------------------------
    # 3. CREATE SPECIFIC WORKSHOP FILTER TABS
    # -------------------------------------------------------------
    lab_tabs = [
        ("Lab - AFM", "Atomic Force Microscopy"),
        ("Lab - Nanomedicine", "Nanomedicine"),
        ("Lab - Quantum Well", "Quantum Well"),
        ("Lab - Nanowires", "Nanowires"),
        ("Lab - HERITAGE-LAB", "HERITAGE"),
        ("Lab - Light & Photonics", "Photonics"),
        ("Lab - Computational", "Computational")
    ]
    
    sub_headers = [
        "Full Name",
        "Email Address",
        "Academic Status / Role",
        "Day 2 Laboratory Workshop",
        "Session Attending"
    ]
    
    for tab_title, search_key in lab_tabs:
        ws_lab = wb.create_sheet(title=tab_title)
        
        # Add Header Row
        ws_lab.append(sub_headers)
        for col_num in range(1, len(sub_headers) + 1):
            cell = ws_lab.cell(row=1, column=col_num)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_lab.row_dimensions[1].height = 26
        
        # Add Dynamic FILTER Formula in Cell A2 that pulls live data from 'Risposte del modulo 1'
        # Wrapped with IFERROR because Google Sheets FILTER does not accept if_empty as a 3rd parameter like Excel does
        ws_lab["A2"] = f'=IFERROR(FILTER(\'Risposte del modulo 1\'!$B$2:$F$1000, ISNUMBER(SEARCH("{search_key}", \'Risposte del modulo 1\'!$E$2:$E$1000))), "No registrations yet for this laboratory")'
        ws_lab["A2"].font = Font(name="Calibri", size=11, bold=False, color="0F172A")
        
        # Adjust columns widths
        ws_lab.column_dimensions["A"].width = 28
        ws_lab.column_dimensions["B"].width = 35
        ws_lab.column_dimensions["C"].width = 25
        ws_lab.column_dimensions["D"].width = 45
        ws_lab.column_dimensions["E"].width = 50

    if default_sheet in wb.sheetnames:
        wb.remove(default_sheet)
        
    output_filename = "Sapienza_Conference_Master_Workbook.xlsx"
    wb.save(output_filename)
    print(f"Successfully generated {output_filename} with all 9 tabs and live formulas!")

if __name__ == "__main__":
    create_conference_workbook()
